#------------------------------------------------------#
# Made by CaptainCluster                               #
# https://github.com/captaincluster                    #
#                                                      #
# Writing the follower data into an Excel file.        #
#------------------------------------------------------#
from openpyxl import Workbook 
import os                     
import sys                     

sys.path.insert(1, os.path.join(sys.path[0], '..'))
import modules.json_data    as json_data
import modules.analyze      as analyze
from variables.values       import Values

VALUES_INSTANCE = Values()

def excelWriteDefaults(workSheet) -> None:
    """
    Adding the default components in the Excel file that won't be changed

    Args:
        workSheet (openpyxl): The worksheet where the changes are appended
    """
    workSheet.cell(row=1, column=1).value   = "Usernames" 
    workSheet.cell(row=1, column=2).value   = "Links"

    workSheet.cell(row=1, column=6).value   = "Total Followers"
    workSheet.cell(row=4, column=6).value   = "Total Followed"
    workSheet.cell(row=7, column=6).value   = "Total Unfollowed"

    workSheet.cell(row=10, column=6).value  = "Net Change"

    workSheet.cell(row=1, column=4).value   = "New Followers"
    workSheet.cell(row=1, column=5).value   = "New Unfollowers"

def excelChangeColumnWidth(workSheet, widthList) -> None:
    """Adjusting the lengths of the columns based on the longest data string"""
    
    columnList = ["A", "B", "C", "D"]   #To identify the columns
    for iteration in range(0, (len(columnList)-1)):
        workSheet.column_dimensions[columnList[iteration]].width = widthList[iteration] + 2
        
def setUpExcel() -> Workbook:
    """ Setting up the workbook and the worksheet

    Returns:
        wb: the Excel workbook
        worksheet: the Excel worksheet
    """
    wb = Workbook()
    workSheet = wb.worksheets[0]
    workSheet.title = VALUES_INSTANCE.EXCEL_WORKSHEET_TITLE

    return wb, workSheet

def excelWritingProcess(followerData) -> None:
    """
    Writing the data into the Excel file

    Args:
        followerData(dict): The data that will be written
    """
    try:
        followerUsernames = followerData["content"][0]["usernames"]
        followerUrls = followerData["content"][0]["urls"]
        widthList = []

        changedFollowershipList = analyze.compareFollowerLists() #Getting the data for fresh followers and those who unfollowed
        followedList, unfollowedList = changedFollowershipList 

        wb, workSheet = setUpExcel()        #Setting up the workbook and a worksheet
        excelWriteDefaults(workSheet)       #Adding the default attributes to the worksheet

        widthList.append(writeListToExcel(workSheet, followerUsernames, VALUES_INSTANCE.USERNAME_COLUMN, None))
        widthList.append(writeListToExcel(workSheet, followerUrls, 2, None))
        widthList.append(writeListToExcel(workSheet, followedList, 3, None))
        widthList.append(writeListToExcel(workSheet, unfollowedList, 4, None))

        workSheet.cell(row=2, column=6).value = len(followerUsernames)
        workSheet.cell(row=5, column=6).value = len(followedList)
        workSheet.cell(row=8, column=6).value = len(unfollowedList)
        workSheet.cell(row=11, column=6).value = (len(followedList)-len(unfollowedList))

        excelChangeColumnWidth(workSheet, widthList)

        wb.save(VALUES_INSTANCE.EXCEL_FILE)    
    except Exception as exception:
        print(VALUES_INSTANCE.EXCEPTION_DEFAULT + str(exception))
    
def writeListToExcel(workSheet, dataList, writeColumn, dataString) -> int:
    """Writing a list, using a for-loop, to an Excel file

    Args:
        workSheet (openpyxl): The worksheet where the changes are appended
        dataList (list): All the data to be written
        writeColumn (int): The column in the Excel file where the data is written
        dataString (string/None): A potential string in front of the given data

    Returns:
        longestLength (int): Integer for determining column width
    """
    longestLength = VALUES_INSTANCE.USERNAMES_DEFAULT_WIDTH
    
    for i, value in enumerate(dataList, start=VALUES_INSTANCE.STARTING_ROW):

        #Handling undefined values
        if value == "":
                value = VALUES_INSTANCE.EXCEL_UNDEFINED_VALUE
        
        #A potential string in front of the written data
        if dataString:
            value = dataString + value

        #The longest length allows us to determine a proper width for a column in Excel
        if len(value) > longestLength:
            longestLength = len(value)

        workSheet.cell(row=i, column=writeColumn).value = value
    
    return longestLength

def writeFollowerData() -> None:
    """The core function for writing the data into an Excel file"""
    followerData = json_data.getJsonData(VALUES_INSTANCE.FOLLOWERDATA_FILE_NAME)
    excelWritingProcess(followerData)
    print(VALUES_INSTANCE.NOTIFY_WRITING_SUCCESSFUL)


if __name__ == "__main__":
    writeFollowerData()
